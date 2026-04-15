"""
Reads giskard_out_stereotypes/scan_summary.json and exports the issues/examples
into an Excel workbook where each unique issue description gets its own sheet.

Columns per sheet:
  reason | input_text | user_command | agent | suggestions
"""

import argparse
import json
import re
import textwrap
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── helpers ────────────────────────────────────────────────────────────────────
COLUMNS = [
    ("reason",       "Reason"),
    ("input_text",   "Input Text"),
    ("user_command", "User Command"),
    ("agent",        "Agent"),
    ("suggestions",  "Suggestions"),
]

HEADER_FILL   = PatternFill("solid", fgColor="4472C4")
HEADER_FONT   = Font(bold=True, color="FFFFFF")
WRAP_ALIGN    = Alignment(wrap_text=True, vertical="top")
COL_WIDTHS    = {
    "reason":       45,
    "input_text":   40,
    "user_command": 30,
    "agent":        55,
    "suggestions":  60,
}


_COMMON_PREFIX = "The model does not satisfy the following requirement: "


def sanitise_sheet_name(name: str, max_len: int = 31) -> str:
    """Excel sheet names: ≤31 chars, no special chars.

    Strips the repetitive Giskard prefix so the unique part of each
    description is used as the tab label.
    """
    if name.startswith(_COMMON_PREFIX):
        name = name[len(_COMMON_PREFIX):]
    name = re.sub(r"[\\/*?\[\]:]", " ", name).strip()
    return textwrap.shorten(name, width=max_len, placeholder="…")


def format_suggestions(suggestions) -> str:
    """Render the suggestions list as a readable multi-line string."""
    if not suggestions:
        return ""
    parts = []
    for i, s in enumerate(suggestions, 1):
        lines = []
        if isinstance(s, dict):
            if s.get("explanation"):
                lines.append(f"[{i}] Explanation: {s['explanation']}")
            if s.get("original_text"):
                lines.append(f"    Original : {s['original_text']}")
            if s.get("transformed_text"):
                lines.append(f"    Suggested: {s['transformed_text']}")
        else:
            lines.append(f"[{i}] {s}")
        parts.append("\n".join(lines))
    return "\n\n".join(parts)


# ── main ───────────────────────────────────────────────────────────────────────
def main():
    base_dir = Path(__file__).parent
    default_input  = base_dir / "giskard_out_stereotypes" / "scan_summary.json"
    default_output = base_dir / "giskard_out_stereotypes" / "scan_summary.xlsx"

    parser = argparse.ArgumentParser(
        description="Convert a Giskard scan_summary.json into an Excel workbook."
    )
    parser.add_argument(
        "-i", "--input",
        type=Path,
        default=default_input,
        metavar="INPUT_JSON",
        help=f"Path to scan_summary.json (default: {default_input})",
    )
    parser.add_argument(
        "-o", "--output",
        type=Path,
        default=None,
        metavar="OUTPUT_XLSX",
        help="Path for the output .xlsx file (default: <input_dir>/<input_stem>.xlsx)",
    )
    args = parser.parse_args()

    input_file: Path  = args.input
    output_file: Path = args.output if args.output else input_file.with_suffix(".xlsx")

    with input_file.open(encoding="utf-8") as fh:
        data = json.load(fh)

    issues = data.get("issues", [])
    if not issues:
        print("No issues found in the JSON file.")
        return

    # Group examples by description (preserving order)
    groups: dict[str, list[dict]] = {}
    for issue in issues:
        description = issue.get("description", "No description")
        examples    = issue.get("examples", [])
        if description not in groups:
            groups[description] = []
        groups[description].extend(examples)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)          # remove the default blank sheet

    # Build a unique sheet name for each description
    used_names: dict[str, int] = {}
    for description, examples in groups.items():
        base_name   = sanitise_sheet_name(description)
        # de-duplicate if two descriptions truncate to the same name
        count       = used_names.get(base_name, 0)
        used_names[base_name] = count + 1
        sheet_name  = base_name if count == 0 else f"{base_name[:28]}_{count}"

        ws = wb.create_sheet(title=sheet_name)

        # ── description row (row 1) ──────────────────────────────────────────
        desc_cell           = ws.cell(row=1, column=1, value=description)
        desc_cell.font      = Font(bold=True, size=12)
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
        desc_cell.fill      = PatternFill("solid", fgColor="D9E1F2")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
        ws.row_dimensions[1].height = 48

        # ── header row (row 2) ──────────────────────────────────────────────
        for col_idx, (_, header_label) in enumerate(COLUMNS, start=1):
            cell            = ws.cell(row=2, column=col_idx, value=header_label)
            cell.font       = HEADER_FONT
            cell.fill       = HEADER_FILL
            cell.alignment  = WRAP_ALIGN

        # ── data rows ───────────────────────────────��───────────────────────
        for row_idx, example in enumerate(examples, start=3):
            system = example.get("system", {})

            row_data = {
                "reason":       example.get("reason", ""),
                "input_text":   system.get("input_text", ""),
                "user_command": system.get("user_command", ""),
                "agent":        system.get("agent", ""),
                "suggestions":  format_suggestions(system.get("suggestions", [])),
            }

            for col_idx, (field_key, _) in enumerate(COLUMNS, start=1):
                cell           = ws.cell(row=row_idx, column=col_idx, value=row_data[field_key])
                cell.alignment = WRAP_ALIGN

        # ── column widths ────────────────────────────────────────────────────
        for col_idx, (field_key, _) in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[field_key]

        # freeze the header row
        ws.freeze_panes = "A3"

        print(f"  Sheet '{sheet_name}': {len(examples)} example(s)")

    wb.save(output_file)
    print(f"\nExcel file saved to: {output_file}")


if __name__ == "__main__":
    main()
